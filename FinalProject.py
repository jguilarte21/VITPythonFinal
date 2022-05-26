from traceback import print_tb
import openpyxl
import requests
import json
from openpyxl import load_workbook


#I think I need to change something here to get it working, not sure 
url = "https://api-nba-v1.p.rapidapi.com/seasons"

headers = {
	"X-RapidAPI-Host": "api-nba-v1.p.rapidapi.com",
	"X-RapidAPI-Key": "6ce27da1f5mshb2050d81a60a5e9p1f578djsn49c8c2f18547"
}

response = requests.request("GET", url, headers=headers)

#Print the response back from the api
print(response)

#print the content of response / response.text 
print(response.text)

#Assigns response object to a variable (Json loads)
clean_data = json.loads(response.text)

#Printing out the clean data
print(clean_data)

#test


#Create a workbook 
wb = load_workbook('NBAStats.xlsx')

#OR

#wb = openpyxl.Workbook() ????

#Go to active worksheet
ws = wb.active




















# #Create the Seasons sheet
# seasonSheet = wb.create_sheet('NBA_Seasons')

# #Designate the first column for all the seasons 
# seasonSheet['A1'] = 'Season'

# seasonSheet['B1'] = 'Year'


















